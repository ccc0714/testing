"""
PRN to Excel Converter
Converts MED-PC .PRN files to Excel format matching datachecker.xlsx structure.

Usage:
    python prn_to_excel.py <input.PRN> [output.xlsx]
    python prn_to_excel.py <input.PRN> [output.xlsx] [--sheet SHEETNAME]

If output.xlsx exists, the new data is appended as a new sheet.
If output.xlsx does not exist, a new workbook is created.
"""

import re
import sys
import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── Event code lookup table ──────────────────────────────────────────────────
# Maps the integer prefix of X-array values to (short_code, label) tuples.
# Extend this dict to support additional experiments.
EVENT_CODES = {
    1:   ("1",  "^BLeft"),
    2:   ("2",  "^BRight"),
    3:   ("3",  "^BCentre"),
    4:   ("4",  "^Lfood"),
    5:   ("5",  "^Cfood"),
    6:   ("6",  "^Rfood"),
    7:   ("7",  "^Lnofood"),
    8:   ("8",  "^Cnofood"),
    9:   ("9",  "^Rnofood"),
    11:  ("11", "^BLeftITI"),
    22:  ("22", "^BRightITI"),
    33:  ("33", "^BCentreITI"),
    41:  ("41", "^foodend"),
    51:  ("51", "^Lnopeck"),
    52:  ("52", "^Cnopeck"),
    53:  ("53", "^Rnopeck"),
    61:  ("61", "^LFoodTrial"),
    62:  ("62", "^LNoFoodTrial"),
    63:  ("63", "^CFoodTrial"),
    64:  ("64", "^CNoFoodTrial"),
    65:  ("65", "^RFoodTrial"),
    66:  ("66", "^RNoFoodTrial"),
    71:  ("71", "^ITIend"),
    90:  ("90", "^Foodend"),
    330: ("330", "^BCentreITI"),
    410: ("410", "^foodend"),
    510: ("510", "^Lnopeck"),
    520: ("520", "^Cnopeck"),
    530: ("530", "^Rnopeck"),
    610: ("610", "^LFoodTrial"),
    620: ("620", "^LNoFoodTrial"),
    630: ("630", "^CFoodTrial"),
    640: ("640", "^CNoFoodTrial"),
    650: ("650", "^RFoodTrial"),
    660: ("660", "^RNoFoodTrial"),
    710: ("710", "^ITIend"),
    790: ("790", "^sessionend"),
    990: ("990", "^sessionstart"),
    7900: ("790", "^sessionend"),
    9900: ("990", "^sessionstart"),
}


# ── PRN Parser ───────────────────────────────────────────────────────────────

def parse_prn(filepath):
    """Parse a MED-PC .PRN file and return a structured dict."""
    with open(filepath, "r", errors="replace") as f:
        raw = f.read()

    lines = [l.rstrip("\r\n") for l in raw.splitlines()]

    data = {
        "file": "",
        "start_date": "",
        "end_date": "",
        "subject": "",
        "experiment": "",
        "group": "",
        "box": "",
        "start_time": "",
        "end_time": "",
        "msn": "",
        "C": [],   # list of (index, value)
        "Q": [],
        "X": [],
    }

    section = None
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Header fields
        if stripped.startswith("File:"):
            data["file"] = stripped[5:].strip()
            continue
        if stripped.startswith("Start Date:"):
            data["start_date"] = stripped[11:].strip()
            continue
        if stripped.startswith("End Date:"):
            data["end_date"] = stripped[9:].strip()
            continue
        if stripped.startswith("Subject:"):
            data["subject"] = stripped[8:].strip()
            continue
        if stripped.startswith("Experiment:"):
            data["experiment"] = stripped[11:].strip()
            continue
        if stripped.startswith("Group:"):
            data["group"] = stripped[6:].strip()
            continue
        if stripped.startswith("Box:"):
            data["box"] = stripped[4:].strip()
            continue
        if stripped.startswith("Start Time:"):
            data["start_time"] = stripped[11:].strip()
            continue
        if stripped.startswith("End Time:"):
            data["end_time"] = stripped[9:].strip()
            continue
        if stripped.startswith("MSN:"):
            data["msn"] = stripped[4:].strip()
            continue

        # Section markers
        if re.match(r"^[CQX]:$", stripped):
            section = stripped[0]
            continue

        # Data rows:  "   idx:   value"
        m = re.match(r"^\s*(\d+):\s+([\d\.\-\+Ee]+)\s*$", stripped)
        if m and section in ("C", "Q", "X"):
            idx = int(m.group(1))
            val = float(m.group(2))
            data[section].append((idx, val))
            continue

    return data


# ── Event-code decoder ───────────────────────────────────────────────────────

def decode_x_value(raw_value):
    """
    Given a raw X-array float (e.g. 620000.1), return
    (time_seconds, event_int, short_code, label).

    MED-PC encodes values as:  event_code * 10000 + time_centiseconds/10
    e.g. 620000.1 → event 62 (^LNoFoodTrial), time 0.1 s
         9900000  → event 990 (^sessionstart), time 0 s
         710013.92 → event 71 (^ITIend), time 13.92 s

    We try each known code (largest first to avoid ambiguity).
    """
    # Sort codes descending so larger codes (e.g. 990, 790) are matched
    # before shorter ones (e.g. 9, 7).
    for code in sorted(EVENT_CODES.keys(), reverse=True):
        multiplier = code * 10000
        if abs(raw_value) >= multiplier:
            residual = raw_value - multiplier
            # Sanity: residual should be a small positive time value
            if 0 <= residual < multiplier:
                short, label = EVENT_CODES[code]
                return residual, code, short, label

    return raw_value, None, "", ""


# ── Summary statistics ───────────────────────────────────────────────────────

def compute_summary(x_data):
    """
    Scan decoded X events and count trial types / food outcomes.
    Returns a dict with keys matching the Excel summary block.
    """
    summary = {
        "L_FOOD_TRIAL": 0, "L_NO_FOOD_TRIAL": 0,
        "C_FOOD_TRIAL": 0, "C_NO_FOOD_TRIAL": 0,
        "R_FOOD_TRIAL": 0, "R_NO_FOOD_TRIAL": 0,
        "L_FOOD": 0,       "L_NO_FOOD": 0,
        "C_FOOD": 0,       "C_NO_FOOD": 0,
        "R_FOOD": 0,       "R_NO_FOOD": 0,
        "TRIAL_NUMBER": 0,
    }

    current_trial_type = None

    for _, raw_val in x_data:
        _, prefix, _, label = decode_x_value(raw_val)

        if label == "^LFoodTrial":
            summary["L_FOOD_TRIAL"] += 1
            current_trial_type = "L"
        elif label == "^LNoFoodTrial":
            summary["L_NO_FOOD_TRIAL"] += 1
            current_trial_type = "L_NO"
        elif label == "^CFoodTrial":
            summary["C_FOOD_TRIAL"] += 1
            current_trial_type = "C"
        elif label == "^CNoFoodTrial":
            summary["C_NO_FOOD_TRIAL"] += 1
            current_trial_type = "C_NO"
        elif label == "^RFoodTrial":
            summary["R_FOOD_TRIAL"] += 1
            current_trial_type = "R"
        elif label == "^RNoFoodTrial":
            summary["R_NO_FOOD_TRIAL"] += 1
            current_trial_type = "R_NO"
        elif label in ("^Lfood", "^foodend") and current_trial_type == "L":
            if label == "^Lfood":
                summary["L_FOOD"] += 1
        elif label == "^Lnofood" and current_trial_type in ("L", "L_NO"):
            if current_trial_type == "L_NO":
                summary["L_NO_FOOD"] += 1
            else:
                summary["L_NO_FOOD"] += 1
        elif label == "^Cfood" and current_trial_type == "C":
            summary["C_FOOD"] += 1
        elif label == "^Cnofood" and current_trial_type in ("C", "C_NO"):
            summary["C_NO_FOOD"] += 1
        elif label == "^Rfood" and current_trial_type == "R":
            summary["R_FOOD"] += 1
        elif label == "^Rnofood" and current_trial_type in ("R", "R_NO"):
            summary["R_NO_FOOD"] += 1
        elif label == "^ITIend":
            current_trial_type = None

    # Count total trials (food + no-food for each side)
    l_total = summary["L_FOOD_TRIAL"] + summary["L_NO_FOOD_TRIAL"]
    c_total = summary["C_FOOD_TRIAL"] + summary["C_NO_FOOD_TRIAL"]
    r_total = summary["R_FOOD_TRIAL"] + summary["R_NO_FOOD_TRIAL"]
    summary["TRIAL_NUMBER"] = l_total + c_total + r_total

    # Proportions
    summary["PROP_L"] = (summary["L_FOOD_TRIAL"] / l_total) if l_total > 0 else 0
    summary["PROP_C"] = (summary["C_FOOD_TRIAL"] / c_total) if c_total > 0 else 0
    summary["PROP_R"] = (summary["R_FOOD_TRIAL"] / r_total) if r_total > 0 else 0

    return summary

# ── Excel writer ─────────────────────────────────────────────────────────────

def write_sheet(ws, data):
    """
    Write parsed PRN data to an openpyxl worksheet,
    matching the layout of datachecker.xlsx.
    """
    # Fonts / styles
    bold = Font(bold=True, name="Arial", size=10)
    normal = Font(name="Arial", size=10)

    def w(row, col, value, font=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font or normal
        if align:
            cell.alignment = Alignment(horizontal=align)
        return cell

    # ── Column widths (approximate match) ──
    col_widths = {1: 12, 2: 16, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14,
                  8: 18, 9: 10, 10: 8, 11: 8, 12: 18, 13: 10, 14: 8,
                  15: 8, 16: 28, 17: 18}
    for c, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = width

    # ── Row 1: File path ──
    w(1, 1, "File:")
    w(1, 2, data["file"])

    # ── Right-side code legend (columns H–L) ──
    legend = [
        (1,  "^BLeft",      10000),
        (2,  "^BRight",     20000),
        (3,  "^BCentre",    30000),
        (4,  "^Lfood",      40000),
        (5,  "^Cfood",      50000),
        (6,  "^Rfood",      60000),
        (7,  "^Lnofood",    70000),
        (8,  "^Cnofood",    80000),
        (9,  "^Rnofood",    90000),
        (61, "^LFoodTrial", 610000),
        (62, "^LNoFoodTrial",620000),
        (63, "^CFoodTrial", 630000),
        (64, "^CNoFoodTrial",640000),
        (65, "^RFoodTrial", 650000),
        (66, "^RNoFoodTrial",660000),
        (41, "^foodend",    410000),
        (51, "^Lnopeck",    510000),
        (52, "^Cnopeck",    520000),
        (53, "^Rnopeck",    530000),
        (71, "^ITIend",     710000),
        (990,"^sessionstart",9900000),
        (790,"^sessionend", 7900000),
    ]
    for i, (code, name, val) in enumerate(legend):
        r = i + 1
        w(r, 8, code)
        w(r, 9, name)
        w(r, 10, "=")
        w(r, 11, val)

    # ── Header block (rows 3–11) ──
    header_rows = [
        (3,  "Start Date:", data["start_date"]),
        (4,  "End Date:",   data["end_date"]),
        (5,  "Subject:",    data["subject"]),
        (6,  "Experiment:", data["experiment"]),
        (7,  "Group:",      data["group"]),
        (8,  "Box:",        data["box"]),
        (9,  "Start Time:", data["start_time"]),
        (10, "End Time:",   data["end_time"]),
        (11, "MSN:",        data["msn"]),
    ]
    for row, label, value in header_rows:
        w(row, 1, label)
        w(row, 2, value)

    # ── C array ──
    current_row = 12
    w(current_row, 1, "C:")
    current_row += 1

    for i, (idx, val) in enumerate(data["C"]):
        # Time index (fraction of day)
        time_frac = datetime.time(hour=int(idx))
        # Store raw index / 24 as fraction (matching observed pattern)
        w(current_row, 2, f"{idx}:00")
        w(current_row, 3, val)
        current_row += 1

    # ── Q array ──
    q_start_row = current_row
    w(q_start_row, 1, "Q:")
    q_start_row += 1

    for i, (idx, val) in enumerate(data["Q"]):
        r = q_start_row + i
        w(r, 2, f"{idx}:00")
        w(r, 3, val)

    q_end_row = q_start_row + len(data["Q"])

    # ── X array + summary block ──
    w(q_end_row, 1, "X:")
    x_data_start = q_end_row + 1

    # Compute summary before writing
    summary = compute_summary(data["X"])

    # Summary labels and values (columns H–Q, rows x_data_start to x_data_start+8)
    sum_rows = [
        ("L FOOD TRIAL",   summary["L_FOOD_TRIAL"],
         "L FOOD",         summary["L_FOOD"],
         "Prop L Trials ending in food", summary["PROP_L"]),
        ("L NO FOOD TRIAL",summary["L_NO_FOOD_TRIAL"],
         "L NO FOOD",      summary["L_NO_FOOD"],
         None, None),
        ("C FOOD TRIAL",   summary["C_FOOD_TRIAL"],
         "C FOOD",         summary["C_FOOD"],
         "Prop C Trials ending in food", summary["PROP_C"]),
        ("C NO FOOD TRIAL",summary["C_NO_FOOD_TRIAL"],
         "C NO FOOD",      summary["C_NO_FOOD"],
         None, None),
        ("R FOOD TRIAL",   summary["R_FOOD_TRIAL"],
         "R FOOD",         summary["R_FOOD"],
         "Prop R Trials ending in food", summary["PROP_R"]),
        ("R NO FOOD TRIAL",summary["R_NO_FOOD_TRIAL"],
         "R NO FOOD",      summary["R_NO_FOOD"],
         None, None),
    ]

    for i, row_data in enumerate(sum_rows):
        r = x_data_start + i
        label1, val1, label2, val2, label3, val3 = row_data
        w(r, 9, label1)
        w(r, 10, val1)
        w(r, 13, label2)
        w(r, 14, val2)
        if label3:
            w(r, 17, label3)
            w(r, 18, val3)

    # Trial number
    r_trial = x_data_start + 7
    w(r_trial, 9, "Trial Number = ")
    w(r_trial, 10, summary["TRIAL_NUMBER"])

    # ── X data rows ──
    for i, (idx, raw_val) in enumerate(data["X"]):
        r = x_data_start + i
        time_sec, prefix, short, label = decode_x_value(raw_val)

        w(r, 2, f"{idx}:00")   
        w(r, 3, raw_val)       # raw X value
        if short:
            w(r, 4, int(short) if short.isdigit() else short)
        w(r, 5, label)
        #if prefix is not None:
            #w(r, 6, prefix * 10000 if prefix <= 999 else prefix * 10)
        #w(r, 7, time_sec if time_sec != raw_val else "")


# ── Main entry point ─────────────────────────────────────────────────────────

def prn_to_excel(prn_path, xlsx_path=None, sheet_name=None):
    """
    Convert a PRN file and write/append to an Excel workbook.

    Args:
        prn_path:   Path to the .PRN input file.
        xlsx_path:  Output .xlsx path. Defaults to same name as PRN.
        sheet_name: Sheet name. Defaults to the Subject field from PRN.
    """
    prn_path = os.path.abspath(prn_path)
    if not os.path.exists(prn_path):
        raise FileNotFoundError(f"PRN file not found: {prn_path}")

    data = parse_prn(prn_path)

    if xlsx_path is None:
        base = os.path.splitext(prn_path)[0]
        xlsx_path = base + ".xlsx"

    if sheet_name is None:
        sheet_name = data["subject"] or os.path.splitext(os.path.basename(prn_path))[0]

    # Load or create workbook
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Ensure unique sheet name
    base_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name} - {counter}"
        counter += 1

    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, data)

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}  (sheet: '{sheet_name}')")
    return xlsx_path


# ── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert a MED-PC PRN file to Excel format."
    )
    parser.add_argument("prn", help="Input .PRN file path")
    parser.add_argument("xlsx", nargs="?", default=None,
                        help="Output .xlsx file path (optional)")
    parser.add_argument("--sheet", default=None,
                        help="Sheet name (defaults to Subject field)")
    args = parser.parse_args()

    prn_to_excel(args.prn, args.xlsx, args.sheet)